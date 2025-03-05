import pickle
import numpy as np
import time
import matplotlib.pyplot as plt
import openpyxl
import ros2_numpy as rnp
import sensor_msgs 

def serialize_data(data, protocol):
    start_time = time.time()
    buffers = []
    if protocol == 5:
        pickled_data = pickle.dumps(
            data, protocol=protocol, buffer_callback=buffers.append)
    elif protocol == 4:
        pickled_data = pickle.dumps(data, protocol=5)
    elif protocol == 0:
        pickled_data = rnp.msgify(sensor_msgs.msg.Image, data, encoding = "64FC1")
    else:
        raise ValueError("Invalid protocol")
    end_time = time.time()
    # Convert to milliseconds
    return pickled_data, buffers, (end_time - start_time) * 1000


def deserialize_data(pickled_data, protocol, buffers=None):
    start_time = time.time()
    if protocol == 5 and buffers:
        data = pickle.loads(pickled_data, buffers=buffers)
    elif protocol == 4:
        data = pickle.loads(pickled_data)
    elif protocol == 0:
        data = rnp.numpify(pickled_data)
    else:
        raise ValueError("Invalid protocol")
    end_time = time.time()
    return data, (end_time - start_time) * 1000  # Convert to milliseconds

# Adjusted benchmark function


def benchmark(protocol):
    data = np.array([np.ones(655360)])  # 5MB Object
    pickled_data, buffers, serialize_time = serialize_data(data, protocol)
    _, deserialize_time = deserialize_data(pickled_data, protocol, buffers)
    return serialize_time, deserialize_time

# warmup
benchmark(0)
benchmark(4)
benchmark(5)

# Perform benchmarks
serialize_time_0, deserialize_time_0 = benchmark(0)
serialize_time_4, deserialize_time_4 = benchmark(4)
serialize_time_5, deserialize_time_5 = benchmark(5)

# Plotting adjustments
labels = ['ROS2 Numpy', 'In-Band', 'Out-of-Band']
serialize_times = [serialize_time_0, serialize_time_4, serialize_time_5]
deserialize_times = [deserialize_time_0, deserialize_time_4, deserialize_time_5]

x = np.arange(len(labels))
width = 0.2  # Make the bars narrower

object_size_mb = 5
serialize_throughput_0 = object_size_mb / (serialize_times[0] / 1000)  # MB/s
deserialize_throughput_0 = object_size_mb / \
    (deserialize_times[0] / 1000)  # MB/s
serialize_throughput_4 = object_size_mb / (serialize_times[1] / 1000)  # MB/s
deserialize_throughput_4 = object_size_mb / \
    (deserialize_times[1] / 1000)  # MB/s
serialize_throughput_5 = object_size_mb / (serialize_times[2] / 1000)  # MB/s
deserialize_throughput_5 = object_size_mb / \
    (deserialize_times[2] / 1000)  # MB/s


throughput_serialize = [serialize_throughput_0, serialize_throughput_4, serialize_throughput_5]
throughput_deserialize = [deserialize_throughput_0, deserialize_throughput_4, deserialize_throughput_5]

# Save throughput values to Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Throughput"

# Write headers
sheet.cell(row=1, column=1, value="Protocol")
sheet.cell(row=1, column=2, value="Serialize Throughput (MB/s)")
sheet.cell(row=1, column=3, value="Deserialize Throughput (MB/s)")

# Write data
for i, label in enumerate(labels, start=2):
    sheet.cell(row=i, column=1, value=label)
    sheet.cell(row=i, column=2, value=throughput_serialize[i-2])
    sheet.cell(row=i, column=3, value=throughput_deserialize[i-2])

# Save the workbook
wb.save("./logs/benchmarks/serialization.xlsx")